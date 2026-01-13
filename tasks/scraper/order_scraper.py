import logging
import time
import re
import os
import glob
import requests
from bs4 import BeautifulSoup
from typing import List, Dict, Any, Optional
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from app.core.config import settings


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class OrderScraper:
    def __init__(self):
        self.BASE_URL = "https://ecosistema.gopharma.com.ve"
        self.LOGIN_URL = f"{self.BASE_URL}/login/admin"
        self.driver = None
        self.download_dir = "/tmp/downloads" # Definido centralmente

    def setup_driver(self):
        if self.driver: return
        
        chrome_options = Options()
        chrome_options.add_argument("--headless=new") 
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        
        # --- SOLUCIÓN AL "PAGE EXPIRED" ---
        # Esto le dice a Chrome: "Si sale una alerta, acéptala y sigue"
        chrome_options.add_argument("--unhandled-alert-behavior=accept")
        
        # Configuración experimental
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        if not os.path.exists(self.download_dir):
            os.makedirs(self.download_dir, mode=0o777)

        prefs = {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_settings.popups": 0,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        service = Service()
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""
        })
        try:
            self.driver.execute_cdp_cmd('Page.setDownloadBehavior', {'behavior': 'allow', 'downloadPath': self.download_dir})
        except: pass

    def login(self):
        if self.driver:
            try:
                # Si estamos en dashboard o admin, es válido
                if "dashboard" in self.driver.current_url or "/admin" in self.driver.current_url: 
                    return True
            except: pass
        else:
            self.setup_driver()
        
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                logger.info(f"🔑 Intento de Login {attempt}/{max_retries}...")
                
                # LIMPIEZA PREVENTIVA: Borrar cookies para evitar "Session Expired"
                try: self.driver.delete_all_cookies()
                except: pass

                self.driver.get(self.LOGIN_URL)
                
                # MANEJO EXPLÍCITO DE ALERTAS (Doble seguridad)
                try:
                    WebDriverWait(self.driver, 3).until(EC.alert_is_present())
                    alert = self.driver.switch_to.alert
                    logger.warning(f"⚠️ Alerta detectada y aceptada: {alert.text}")
                    alert.accept()
                except: pass

                # Validar si ya entró
                if "dashboard" in self.driver.current_url: return True

                wait = WebDriverWait(self.driver, 20)
                email_input = wait.until(EC.presence_of_element_located((By.NAME, "email")))
                email_input.clear()
                email_input.send_keys(settings.GOPHARMA_EMAIL)
                time.sleep(0.5)
                
                pass_input = self.driver.find_element(By.NAME, "password")
                pass_input.clear()
                pass_input.send_keys(settings.GOPHARMA_PASSWORD)
                pass_input.send_keys(Keys.RETURN)
                
                # Esperamos dashboard O admin
                wait.until(lambda d: "dashboard" in d.current_url or "/admin" in d.current_url)
                logger.info("✅ Login Exitoso.")
                return True

            except Exception as e:
                logger.warning(f"⚠️ Falló intento {attempt}: {e}")
                # Si hay una alerta bloqueando, intentamos aceptarla de nuevo
                try: self.driver.switch_to.alert.accept()
                except: pass
                
                if self.driver:
                    self.driver.save_screenshot(f"/app/static/error_login_try_{attempt}.png")
                
                time.sleep(2)
        
        logger.error("❌ Se agotaron los intentos de Login.")
        self.close_driver()
        return False

    def close_driver(self):
        if self.driver:
            try: self.driver.quit()
            except: pass
            self.driver = None

    def download_official_excel(self, order_id: str):
        if not self.login(): return None, None
        
        # 1. Limpieza
        for f in glob.glob(os.path.join(self.download_dir, "*")):
            try: os.remove(f)
            except: pass

        # REFUERZO DE PERMISOS
        try:
            self.driver.execute_cdp_cmd('Page.setDownloadBehavior', {'behavior': 'allow', 'downloadPath': self.download_dir})
        except: pass

        logger.info(f"🤖 Robot: Buscando CSV para #{order_id}...")
        
        try:
            # Navegar a la lista
            self.driver.get(f"{self.BASE_URL}/admin/order/list/all")
            
            # 2. FILTRADO
            search_input = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, "datatableSearch_")))
            search_input.clear()
            search_input.send_keys(order_id)
            search_input.send_keys(Keys.RETURN)
            time.sleep(3) 

            # 3. INTERACCIÓN QUIRÚRGICA
            try:
                # Abrir Dropdown
                export_btn = self.driver.find_element(By.CSS_SELECTOR, ".js-hs-unfold-invoker")
                self.driver.execute_script("arguments[0].click();", export_btn)
                time.sleep(1)
                
                # Encontrar botón CSV
                csv_btn = self.driver.find_element(By.XPATH, "//a[contains(@id, 'export-csv') or contains(text(), 'CSV')]")
                
                # --- EL TRUCO MAESTRO (DOM SANITIZATION) ---
                # Removemos target="_blank" para obligar descarga en la misma pestaña
                # Esto evita el bloqueo "about:blank#blocked"
                self.driver.execute_script("arguments[0].removeAttribute('target');", csv_btn)
                logger.info("💉 JS: Atributo target='_blank' eliminado.")
                
                # Click nativo
                self.driver.execute_script("arguments[0].click();", csv_btn)
                logger.info("✅ Clic en CSV realizado.")
                
            except Exception as e:
                logger.error(f"❌ Falló clic UI: {e}")
                self.driver.save_screenshot("/app/static/error_menu_click.png")
                return None, None
            
            # 4. ESPERA ACTIVA
            file_path = None
            for i in range(40):
                files = os.listdir(self.download_dir)
                if i % 5 == 0: logger.info(f"⏳ [{i}s] Carpeta: {files}")
                
                candidates = [f for f in files if f.endswith(".csv")]
                if candidates:
                    full_path = os.path.join(self.download_dir, candidates[0])
                    if os.path.getsize(full_path) > 50:
                        file_path = full_path
                        break
                time.sleep(1)
            
            if not file_path:
                logger.error("❌ Timeout descarga CSV.")
                self.driver.save_screenshot("/app/static/error_timeout.png")
                return None, None

            # 5. CONVERSIÓN A EXCEL "PRETTY" (FORMATO FICHA)
            logger.info(f"📄 Maquetando Excel para ATC: {file_path}")
            try:
                import csv
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from io import BytesIO

                # Configuración de Estilos
                BOLD_FONT = Font(bold=True, name='Arial', size=10)
                TITLE_FONT = Font(bold=True, name='Arial', size=14, color="FFFFFF")
                HEADER_FILL = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid") # Azul GoAnalisis
                SUBHEADER_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid") # Gris suave
                CENTER = Alignment(horizontal='center', vertical='center')
                LEFT = Alignment(horizontal='left', vertical='center')
                THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Ficha {order_id}"
                
                # Desactivar cuadrícula para look de "Reporte"
                ws.sheet_view.showGridLines = False

                # Leer datos del CSV (Asumimos que es 1 sola fila de datos + headers)
                data = {}
                with open(file_path, 'r', encoding='utf-8-sig') as f: # utf-8-sig quita el BOM
                    reader = csv.DictReader(f)
                    for row in reader:
                        data = row # Tomamos la primera (y única) fila
                        break
                
                if not data:
                    raise Exception("El CSV descargado está vacío")

                # --- SECCIÓN 1: ENCABEZADO ---
                ws.merge_cells('B2:E2')
                cell = ws['B2']
                cell.value = f"FICHA DE PEDIDO #{data.get('ID del pedido', order_id)}"
                cell.font = TITLE_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER

                ws['F2'].value = data.get('Estado del pedido', 'N/A').upper()
                ws['F2'].font = Font(bold=True, color="FF0000" if 'Cancelado' in data.get('Estado del pedido','') else "008000")
                ws['F2'].alignment = CENTER
                ws['F2'].border = THIN_BORDER

                # --- SECCIÓN 2: DATOS DEL CLIENTE (Izquierda) Y LOGÍSTICA (Derecha) ---
                # Fila 4
                ws['B4'].value = "CLIENTE:"
                ws['B4'].font = BOLD_FONT
                ws['C4'].value = data.get('Nombre del cliente', 'N/A')
                
                ws['E4'].value = "FECHA:"
                ws['E4'].font = BOLD_FONT
                ws['F4'].value = data.get('Fecha', 'N/A')

                # Fila 5
                ws['B5'].value = "TELÉFONO:"
                ws['B5'].font = BOLD_FONT
                ws['C5'].value = data.get('Teléfono del cliente', 'N/A')

                ws['E5'].value = "TIENDA:"
                ws['E5'].font = BOLD_FONT
                ws['F5'].value = data.get('Nombre de la tienda', 'N/A')

                # Fila 6
                ws['B6'].value = "EMAIL:"
                ws['B6'].font = BOLD_FONT
                ws['C6'].value = data.get('Correo electrónico del cliente', 'N/A')

                ws['E6'].value = "TIPO:"
                ws['E6'].font = BOLD_FONT
                ws['F6'].value = data.get('Tipo de pedido', 'N/A')

                # --- SECCIÓN 3: TABLA FINANCIERA (Crucial para ATC) ---
                ws.merge_cells('B8:F8')
                ws['B8'].value = "DETALLE FINANCIERO"
                ws['B8'].font = BOLD_FONT
                ws['B8'].fill = SUBHEADER_FILL
                ws['B8'].alignment = CENTER

                # Headers Tabla
                headers = [("CONCEPTO", "B9"), ("USD ($)", "C9"), ("VED (Bs)", "D9")]
                for text, pos in headers:
                    ws[pos].value = text
                    ws[pos].font = BOLD_FONT
                    ws[pos].border = THIN_BORDER
                    ws[pos].alignment = CENTER

                # Filas de montos
                financials = [
                    ("Subtotal Productos", data.get('Precio del artículo (USD)', '0'), data.get('Precio del artículo (VED)', '0')),
                    ("Delivery (Envío)", data.get('Cargo de entrega (USD)', '0'), data.get('Cargo de entrega (VED)', '0')),
                    ("Tarifa Servicio", data.get('Tarifa de servicio (USD)', '0'), data.get('Tarifa de servicio (VED)', '0')),
                    ("Impuestos", data.get('Impuesto (USD)', '0'), data.get('Impuesto (VED)', '0')),
                    ("Descuentos", f"-{data.get('Monto descontado (USD)', '0')}", f"-{data.get('Monto descontado (VED)', '0')}")
                ]

                curr_row = 10
                for label, usd, ved in financials:
                    ws[f'B{curr_row}'].value = label
                    ws[f'B{curr_row}'].border = THIN_BORDER
                    
                    # Convertir a float para que Excel lo reconozca como número
                    try: ws[f'C{curr_row}'].value = float(usd.replace(',','.'))
                    except: ws[f'C{curr_row}'].value = usd
                    
                    try: ws[f'D{curr_row}'].value = float(ved.replace(',','.'))
                    except: ws[f'D{curr_row}'].value = ved
                    
                    # Estilo Moneda
                    ws[f'C{curr_row}'].number_format = '"$"#,##0.00'
                    ws[f'D{curr_row}'].number_format = '"Bs."#,##0.00'
                    
                    ws[f'C{curr_row}'].border = THIN_BORDER
                    ws[f'D{curr_row}'].border = THIN_BORDER
                    curr_row += 1

                # TOTAL FINAL
                ws[f'B{curr_row}'].value = "MONTO TOTAL"
                ws[f'B{curr_row}'].font = BOLD_FONT
                ws[f'B{curr_row}'].fill = SUBHEADER_FILL
                ws[f'B{curr_row}'].border = THIN_BORDER
                
                try: ws[f'C{curr_row}'].value = float(data.get('Monto total (USD)', '0').replace(',','.'))
                except: ws[f'C{curr_row}'].value = 0
                ws[f'C{curr_row}'].font = BOLD_FONT
                ws[f'C{curr_row}'].fill = SUBHEADER_FILL
                ws[f'C{curr_row}'].border = THIN_BORDER
                ws[f'C{curr_row}'].number_format = '"$"#,##0.00'
                
                try: ws[f'D{curr_row}'].value = float(data.get('Monto total (VED)', '0').replace(',','.'))
                except: ws[f'D{curr_row}'].value = 0
                ws[f'D{curr_row}'].font = BOLD_FONT
                ws[f'D{curr_row}'].fill = SUBHEADER_FILL
                ws[f'D{curr_row}'].border = THIN_BORDER
                ws[f'D{curr_row}'].number_format = '"Bs."#,##0.00'

                # --- SECCIÓN 4: DATOS DE PAGO (Derecha de Finanzas) ---
                # Usamos el espacio a la derecha (Columna F)
                start_pay = 9
                ws[f'F{start_pay}'].value = "DATOS DE PAGO"
                ws[f'F{start_pay}'].font = BOLD_FONT
                ws[f'F{start_pay}'].fill = SUBHEADER_FILL
                ws[f'F{start_pay}'].alignment = CENTER
                ws[f'F{start_pay}'].border = THIN_BORDER

                ws[f'E{start_pay+1}'].value = "Método:"
                ws[f'F{start_pay+1}'].value = data.get('Método de pago', 'N/A')
                ws[f'F{start_pay+1}'].font = Font(bold=True, color="0000FF")

                ws[f'E{start_pay+2}'].value = "Referencia:"
                ws[f'F{start_pay+2}'].value = data.get('Referencia', 'N/A')
                
                ws[f'E{start_pay+3}'].value = "Estado Pago:"
                ws[f'F{start_pay+3}'].value = data.get('Estado del pago', 'N/A')
                status_color = "008000" if data.get('Estado del pago') == 'Pagado' else "FF0000"
                ws[f'F{start_pay+3}'].font = Font(bold=True, color=status_color)

                ws[f'E{start_pay+4}'].value = "Tasa Cambio:"
                ws[f'F{start_pay+4}'].value = data.get('Tasa de cambio', 'N/A')

                # AJUSTE DE ANCHO DE COLUMNAS
                ws.column_dimensions['A'].width = 2
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 25

                # Guardamos en memoria
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                logger.info("✨ Excel 'Pretty' generado con éxito.")
                return output.read(), f"Ficha_Pedido_{order_id}.xlsx"

            except ImportError:
                # Fallback CSV
                with open(file_path, "rb") as f: content = f.read()
                return content, f"Orden_{order_id}.csv"
            except Exception as e:
                logger.error(f"⚠️ Error maquetando Excel: {e}")
                # Si falla el maquetado, entregamos el CSV crudo por seguridad
                with open(file_path, "rb") as f: content = f.read()
                return content, f"Orden_{order_id}.csv"

        except Exception as e:
            logger.error(f"❌ Crash Scraper: {e}")
            self.driver.save_screenshot("/app/static/error_crash.png")
            return None, None
        finally:
            self.close_driver()

    def _parse_duration(self, row_element) -> str:
        """Extrae el texto de duración de la fila."""
        try:
            # Buscamos en la segunda columna (td[2]) que suele tener la fecha y duración
            # Ojo: XPath relativo a la fila
            div = row_element.find_element(By.XPATH, ".//div[contains(., 'Duración de tiempo')]")
            return " ".join(div.text.replace("Duración de tiempo:", "").strip().split())
        except: return ""

    def get_recent_order_ids(self, limit: int = 25) -> List[Dict[str, str]]:
        """
        Retorna lista de dicts: [{'id': '123', 'duration': '1h 5m'}]
        """
        if not self.driver: self.setup_driver(); self.login()
        orders_found = []
        
        try:
            self.driver.get(self.orders_url)
            WebDriverWait(self.driver, 15).until(EC.presence_of_element_located((By.ID, "datatable")))
            
            rows = self.driver.find_elements(By.XPATH, "//table[@id='datatable']/tbody/tr")
            
            for row in rows:
                if len(orders_found) >= limit: break
                try:
                    # Ignorar filas basura
                    if "Carrito" in row.text or "group-separator" in row.get_attribute("class"): continue

                    # ID
                    link = row.find_element(By.XPATH, ".//a[contains(@href, '/order/details/')]")
                    href = link.get_attribute("href")
                    order_id = href.split("/")[-1]
                    
                    # Duración (La rescatamos aquí)
                    duration = self._parse_duration(row)

                    if order_id.isdigit():
                        orders_found.append({"id": order_id, "duration": duration})
                except: continue
                
        except Exception as e:
            logger.error(f"Error get_recent: {e}")
        
        return orders_found

    def get_historical_ids(self, max_pages: int = None) -> List[Dict[str, str]]:
        """
        Navega por la paginación hasta el final.
        Si max_pages es None, sigue hasta que no haya botón 'Siguiente'.
        """
        if not self.driver: self.setup_driver(); self.login()
        all_data = []
        
        try:
            self.driver.get(self.orders_url)
            WebDriverWait(self.driver, 15).until(EC.presence_of_element_located((By.ID, "datatable")))
            
            current_page = 1
            while True:
                # Freno de emergencia opcional (si se pasa un número explícito)
                if max_pages and current_page > max_pages:
                    logger.info(f"🛑 Límite de seguridad alcanzado ({max_pages} págs). Deteniendo.")
                    break

                logger.info(f"📄 Escaneando pág {current_page}...")
                
                rows = self.driver.find_elements(By.XPATH, "//table[@id='datatable']/tbody/tr")
                page_data = []
                
                for row in rows:
                    try:
                        if "Carrito" in row.text: continue
                        
                        link = row.find_element(By.XPATH, ".//a[contains(@href, '/order/details/')]")
                        order_id = link.get_attribute("href").split("/")[-1]
                        duration = self._parse_duration(row)
                        
                        if order_id.isdigit():
                            page_data.append({"id": order_id, "duration": duration})
                    except: continue
                
                all_data.extend(page_data)
                
                # --- LÓGICA DE PAGINACIÓN INFINITA ---
                try:
                    next_btn = self.driver.find_element(By.XPATH, "//a[@aria-label='Next »']")
                    
                    # Verificamos si el botón está deshabilitado (clase 'disabled' en el padre <li>)
                    parent = next_btn.find_element(By.XPATH, "./..")
                    if "disabled" in parent.get_attribute("class"):
                        logger.info("🚫 Fin de la paginación (Botón deshabilitado).")
                        break
                    
                    # Click para avanzar
                    self.driver.execute_script("arguments[0].click();", next_btn)
                    
                    # Esperamos que cargue la siguiente página
                    # (Pequeña pausa técnica para no saturar y dar tiempo al DOM)
                    time.sleep(2) 
                    
                    current_page += 1
                except NoSuchElementException:
                    logger.info("🚫 No se encontró botón siguiente. Fin de la lista.")
                    break
                except Exception as e:
                    logger.error(f"⚠️ Error al cambiar de página: {e}")
                    break
                    
        except Exception as e:
            logger.error(f"Error backfill: {e}")
        
        # Deduplicar
        unique = {d['id']: d for d in all_data}
        return list(unique.values())
